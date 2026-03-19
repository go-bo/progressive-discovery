#!/usr/bin/env python3
"""Generate an Excel matrix of all Platform Primer page configs."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Primer Config Matrix"

# ── Data extracted from primer-configs.tsx ──────────────────
#
# Dynamism is scored on two independent axes:
#   Axis 1 — Content dynamism:  Does the discovery row need runtime data?
#            (SKU-filtered recipe templates)
#   Axis 2 — CTA dynamism:     Does the CTA need to deep-link into a
#            product creation flow (beyond simple navigation)?
#
# Color system:
#   GREEN  = neither axis (pure static content + navigate-only CTA)
#   YELLOW = one axis (SKU-filtered content OR creation CTA, not both)
#   RED    = both axes (SKU-filtered content AND creation CTA)

pages = [
    {
        "page": "Reports",
        "page_id": "reports",
        "hero_variant": "Action Card",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": True,
        "content_detail": "18 recipe templates filtered by purchased SKU (time, payroll, spend, benefits, IT, talent)",
        "cta_dynamic": True,
        "cta_detail": "\"Create a report\" opens report builder; recipe cards need template IDs to pre-populate",
        "primary_cta": "Create a report",
        "secondary_cta": "Start from template",
    },
    {
        "page": "Data Catalog",
        "page_id": "data-catalog",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Unlock",
        "full_page": True,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Unlock cards and templates are static content",
        "cta_dynamic": False,
        "cta_detail": "\"Open data catalog\" is navigation only — no creation concept",
        "primary_cta": "Open data catalog",
        "secondary_cta": "—",
    },
    {
        "page": "Documents",
        "page_id": "documents",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": True,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Capability features and template recipes are static content",
        "cta_dynamic": True,
        "cta_detail": "\"Create template\" deep-links to template builder; recipe cards (Offer letter, NDA) need template IDs if clickable",
        "primary_cta": "Open Documents",
        "secondary_cta": "Create template",
    },
    {
        "page": "Flow Configuration",
        "page_id": "flow-configuration",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Capability features and unlock cards are static content",
        "cta_dynamic": True,
        "cta_detail": "\"Edit hiring flow\" routes to a specific flow editor tab",
        "primary_cta": "Configure your flows",
        "secondary_cta": "Edit hiring flow",
    },
    {
        "page": "Data Pipelines",
        "page_id": "data-pipelines",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": True,
        "content_detail": "18 recipe templates filtered by purchased SKU (time, payroll, spend, benefits, IT, talent)",
        "cta_dynamic": True,
        "cta_detail": "\"Set up your first pipeline\" opens pipeline wizard; recipe cards need pipeline config IDs",
        "primary_cta": "Set up your first pipeline",
        "secondary_cta": "Start from prebuilt sync",
    },
    {
        "page": "Data Permissions",
        "page_id": "data-permissions",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Template presets are static content, no SKU filtering",
        "cta_dynamic": True,
        "cta_detail": "\"Create a permission set\" opens set builder; preset cards need permission set template IDs",
        "primary_cta": "Create a permission set",
        "secondary_cta": "Start from preset",
    },
    {
        "page": "Workflow Studio",
        "page_id": "workflow-studio",
        "hero_variant": "Action Card",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": True,
        "content_detail": "18 recipe templates filtered by purchased SKU (time, payroll, spend, benefits, IT, talent)",
        "cta_dynamic": True,
        "cta_detail": "\"Build a workflow\" opens workflow builder; recipe cards need workflow template IDs",
        "primary_cta": "Build a workflow",
        "secondary_cta": "Start from recipe",
    },
    {
        "page": "Approvals",
        "page_id": "approvals",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Template recipes are static content, no SKU filtering",
        "cta_dynamic": True,
        "cta_detail": "\"Set up an approval chain\" opens chain builder; template cards need approval chain template IDs",
        "primary_cta": "Set up an approval chain",
        "secondary_cta": "Start from template",
    },
    {
        "page": "Developer",
        "page_id": "developer",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Template",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Starter templates are static content",
        "cta_dynamic": True,
        "cta_detail": "\"Create an app\" opens app registration; starter cards need project scaffold IDs",
        "primary_cta": "Create an app",
        "secondary_cta": "Start from starter",
    },
    {
        "page": "Sandbox",
        "page_id": "sandbox",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": False,
        "has_visual": False,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Capability features and test scenarios are static content",
        "cta_dynamic": False,
        "cta_detail": "\"Set up your sandbox\" triggers provisioning — primary CTA is already the creation action, no secondary needed",
        "primary_cta": "Set up your sandbox",
        "secondary_cta": "—",
    },
    {
        "page": "Activity Log",
        "page_id": "activity-log",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Capability features and unlock cards are static content",
        "cta_dynamic": False,
        "cta_detail": "\"Open activity log\" is navigation only — no creation concept",
        "primary_cta": "Open activity log",
        "secondary_cta": "Configure alerts",
    },
    {
        "page": "Permissions",
        "page_id": "permissions",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Template, Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Capability features, profile templates, and unlock cards are static content",
        "cta_dynamic": True,
        "cta_detail": "\"Create a role\" opens role builder; profile template cards need role template IDs",
        "primary_cta": "Create a role",
        "secondary_cta": "Start from profile template",
    },
    {
        "page": "Saved Supergroups",
        "page_id": "saved-supergroups",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Unlock",
        "full_page": False,
        "has_visual": True,
        "all_variants_configured": "Capability, Unlock",
        "content_dynamic": False,
        "content_detail": "Unlock cards and capability features are static content",
        "cta_dynamic": False,
        "cta_detail": "\"Create a group\" is the primary CTA — straightforward creation, no template pre-population needed",
        "primary_cta": "Create a group",
        "secondary_cta": "—",
    },
    {
        "page": "Chat",
        "page_id": "chat",
        "hero_variant": "Hero Banner",
        "discovery_variant": "Capability",
        "full_page": True,
        "has_visual": True,
        "all_variants_configured": "Capability only",
        "content_dynamic": False,
        "content_detail": "Capability features are static content",
        "cta_dynamic": True,
        "cta_detail": "\"Create a channel\" opens channel creation flow — needs to invoke Chat product's creation UI",
        "primary_cta": "Set up Chat",
        "secondary_cta": "Create a channel",
    },
]

# ── Compute dynamism level per page ──────────────────────────

for p in pages:
    axes = int(p["content_dynamic"]) + int(p["cta_dynamic"])
    if axes == 0:
        p["dynamism_level"] = "Green"
        p["dynamism_label"] = "Static"
        p["dynamism_summary"] = "No dynamic content. No creation deep-link. Pure config + navigation."
    elif axes == 1:
        p["dynamism_level"] = "Yellow"
        axis = "Content" if p["content_dynamic"] else "CTA"
        detail = p["content_detail"] if p["content_dynamic"] else p["cta_detail"]
        p["dynamism_label"] = f"1 axis ({axis})"
        p["dynamism_summary"] = detail
    else:
        p["dynamism_level"] = "Red"
        p["dynamism_label"] = "2 axes (Content + CTA)"
        p["dynamism_summary"] = f"Content: {p['content_detail']}. CTA: {p['cta_detail']}"

# ── Styling ─────────────────────────────────────────────────

header_font = Font(name="Inter", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

body_font = Font(name="Inter", size=11)
bold_font = Font(name="Inter", size=11, bold=True)
body_alignment = Alignment(vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
green_font = Font(name="Inter", size=11, bold=True, color="155724")

yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
yellow_font = Font(name="Inter", size=11, bold=True, color="856404")

red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
red_font = Font(name="Inter", size=11, bold=True, color="721C24")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── Headers ─────────────────────────────────────────────────

headers = [
    "Page",
    "Page ID",
    "Hero Variant",
    "Discovery Row",
    "Full Page?",
    "Primary CTA",
    "Secondary CTA",
    "Dynamism",
    "Content Dynamic?",
    "CTA Dynamic?",
    "Detail",
    "All Configured Variants",
]

col_widths = [22, 20, 15, 15, 12, 26, 26, 20, 16, 16, 65, 28]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 36

# ── Dynamism level color map ────────────────────────────────

level_styles = {
    "Green":  (green_fill, green_font),
    "Yellow": (yellow_fill, yellow_font),
    "Red":    (red_fill, red_font),
}

# ── Data rows ───────────────────────────────────────────────

for row_idx, p in enumerate(pages, 2):
    fill, level_font = level_styles[p["dynamism_level"]]

    values = [
        p["page"],
        p["page_id"],
        p["hero_variant"],
        p["discovery_variant"],
        "Yes" if p["full_page"] else "No",
        p["primary_cta"],
        p["secondary_cta"],
        p["dynamism_label"],
        "Yes" if p["content_dynamic"] else "No",
        "Yes" if p["cta_dynamic"] else "No",
        p["dynamism_summary"],
        p["all_variants_configured"],
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border

        if col_idx == 8:
            cell.fill = fill
            cell.font = level_font
            cell.alignment = center_alignment
        elif col_idx in (9, 10):
            cell.font = bold_font if val == "Yes" else body_font
            cell.fill = fill if val == "Yes" else PatternFill()
            cell.alignment = center_alignment
        elif col_idx == 11:
            cell.font = body_font
            cell.fill = fill
            cell.alignment = body_alignment
        elif col_idx in (3, 4, 5):
            cell.font = body_font
            cell.alignment = center_alignment
        else:
            cell.font = body_font
            cell.alignment = body_alignment

    ws.row_dimensions[row_idx].height = 48 if p["dynamism_level"] == "Red" else 36

# ── Legend sheet ─────────────────────────────────────────────

lg = wb.create_sheet("Legend")
lg.column_dimensions["A"].width = 18
lg.column_dimensions["B"].width = 70

legend_rows = [
    ("Dynamism Level", "Meaning"),
    ("Green — Static", "No dynamic content, no creation deep-link. Primer is pure configuration + navigation. Ship as static config."),
    ("Yellow — 1 Axis", "ONE of: (a) discovery row content filtered by purchased SKUs, OR (b) CTA triggers a product creation flow. Requires one integration point."),
    ("Red — 2 Axes", "BOTH: discovery row content is SKU-filtered AND CTA triggers creation with pre-populated data. Requires integration with both entitlements API and product creation API."),
]

for r, (label, desc) in enumerate(legend_rows, 1):
    a = lg.cell(row=r, column=1, value=label)
    b = lg.cell(row=r, column=2, value=desc)
    a.border = thin_border
    b.border = thin_border
    b.alignment = body_alignment
    if r == 1:
        a.font = header_font
        a.fill = header_fill
        a.alignment = header_alignment
        b.font = header_font
        b.fill = header_fill
        b.alignment = header_alignment
    elif r == 2:
        a.font = green_font
        a.fill = green_fill
        b.font = body_font
        b.fill = green_fill
    elif r == 3:
        a.font = yellow_font
        a.fill = yellow_fill
        b.font = body_font
        b.fill = yellow_fill
    elif r == 4:
        a.font = red_font
        a.fill = red_fill
        b.font = body_font
        b.fill = red_fill

lg.row_dimensions[1].height = 28
lg.row_dimensions[2].height = 36
lg.row_dimensions[3].height = 50
lg.row_dimensions[4].height = 50

# ── Freeze panes + filter ───────────────────────────────────

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(pages) + 1}"

# ── Save ────────────────────────────────────────────────────

output_path = "/Users/paulbest/Documents/htdocs/progressive-discovery/docs/primer-config-matrix.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
